# -*- coding: utf-8 -*-
"""
Created on Thu Mar  8 22:32:11 2018

@author: gregs
"""

### TO DO:

### Improve error handling for empty cells
### Improve screen output

import datetime

### Filenames and parameters defined here
version = "2.5"
sourcefile = "input(2.5-8020).xlsx"
sourcesheet = "data" #the sheet in the sourcefile that you are reading from
rundate = '{date:%Y-%m-%d-%H-%M-%S}'.format( date=datetime.datetime.now())
print(rundate)

outputfilename = "{}-Output(v{}-{}).xlsx".format(sourcesheet,version,rundate)
#rowstoparse = 600  #Should be at least as big as the number of rows in the xlsx. IMSGC.xlsx should be 327095 including the header row at top.

## Printing Basic Info to Screen
print()
print('Coord2gene version',version)
print()


### Read the source data XLSX file

print('Opening up the source XLSX file...', sourcefile)
from openpyxl import load_workbook
wbs = load_workbook(sourcefile, read_only=True)
print ("Opened file", sourcefile)
sheet_ranges = wbs[sourcesheet]
datasheet = wbs.get_sheet_by_name(sourcesheet)

from openpyxl import Workbook
from pyensembl import EnsemblRelease


#Count the number of rows
rowcount = 0
print('Counting the number of rows...')
datasheet.max_row = None
for row in datasheet.iter_rows():
    rowcount += 1

print('Looking up ', rowcount, 'rows from sheet titled', sourcesheet,'in file', sourcefile, 'putting results in',outputfilename)
input('Press Enter to Continue...')


####===========   MAIN LOOP  ============================

print('Working...')
outputfile = Workbook(write_only=True)
outputdatasheet = outputfile.create_sheet(title = 'Output')

header=["Chr ID","Chr Pos","Gene 1","Gene 2","Gene 3","Gene 4"]
outputdatasheet.append(header)

for row in datasheet.iter_rows('A2:B%d' % rowcount):
    coord = [0,0]


    #Parse cells within row (Make list with contig, Chrom. position)
    for cell in row:
        #c=1
        if  cell.value:
            coordidx=cell.column-1
            coord[coordidx]=cell.value
            #c += 1
        else:
            print('Error: Row', row, 'contains invalid coordinate.')

    #Look up Gene Name
    data = EnsemblRelease(75)
    gene_name = data.gene_names_at_locus(contig=coord[0],position=coord[1])

    coord=coord+gene_name
    outputdatasheet.append(coord)


# Save the output file.
outputfile.save(outputfilename)
print()
print ('Wrote data to file',outputfilename)
print()


